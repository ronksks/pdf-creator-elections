import re
import PyPDF2
import os
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook

# before running- you need to fix the input data.
# fix the address in excel with this: =IF(T2="", V2, T2 & " " & U2 & ", " & V2)


# pdfmetrics.registerFont(TTFont('Hebrew', 'Assistant-VariableFont_wght.ttf'))
pdfmetrics.registerFont(TTFont('Hebrew', 'arial-hebrew-bold.ttf'))
pdfmetrics.registerFont(TTFont('Hebrew_david', 'DavidLibre-Bold.ttf'))


# def calculate_text_start(line_start, line_width, font_size, name):
#     # Step 1: Calculate the width of the text
#     # char_width = font_size * 1.5  # This is an estimate
#     # text_width = len(name) * char_width
#     # is 17
#     len_Name= len(name)
#     print('name: ', name)
#     # Step 2: Calculate the starting point for the text
#     text_start = line_start + (line_width - text_width) / 2
#     # returned_value = line_start+ (line_start-text_start)
#
#     return text_start
def calculate_num(num, text, jump, starting_length):
    # Calculate the length of the string
    length = len(text)
    if length == 29:
        return num + 5
    # If the length is 17 characters or more, return the same number
    if length >= starting_length:
        return num

    # For each 2 letters less than 17, return the number -10
    else:
        diff = starting_length - length
        return num + (diff // 2) * jump


# def reverse_slicing(s):
#     if isinstance(s, str):
#         return s[::-1]
#     return str(s)  # Convert non-string values to strings
def reverse_slicing(s):
    if isinstance(s, str):
        result = ""
        non_digit_chars = ""
        digit_chars = ""

        for char in s:
            if char.isdigit():
                if non_digit_chars:
                    result = non_digit_chars[::-1] + result
                    non_digit_chars = ""
                digit_chars += char
            else:
                if digit_chars:
                    result = digit_chars + result
                    digit_chars = ""
                non_digit_chars += char

        result = digit_chars + result
        result = non_digit_chars[::-1] + result

        return result
    return str(s)


# Define a mapping of parameter names/indices to x, y coordinates
# מזכיר/סדרן/אב בית סדרן
parameter_coordinates_rest = {
    # changed height for ramat hasharon
    # "Parameter1": (357, 567),  # שם מלא
    # "Parameter2": (210, 567),  # רחוב+עיר
    # "Parameter3": (100, 567),  # תז
    "Parameter1": (357, 570),  # שם מלא
    "Parameter2": (210, 570),  # רחוב+עיר
    "Parameter3": (100, 570),  # תז
    "Parameter4": (413, 553),  # מספר ריכוז
    # "Parameter4": (413, 550),  # מספר ריכוז
    "Parameter5": (272, 553),  # מיקום קלפי\ ריכוז
    "Parameter6": (148, 553),  # כתובת מיקום קלפי\ ריכוז
    # "Parameter7": (330, 514),  # קלפיות החלפה
    "Parameter7": (413, 553),  # מספר קלפי
    # "Parameter5": (272, 550),  # מיקום קלפי\ ריכוז
    # "Parameter6": (148, 550),  # כתובת מיקום קלפי\ ריכוז
    # # "Parameter7": (330, 514),  # קלפיות החלפה
    # "Parameter7": (413, 550),  # מספר קלפי
    "Parameter8": (0, 0),  # קלפיות החלפה

}
parameter_coordinates_second_or_kolet = {

    "Parameter1": (357, 570),  # שם מלא
    "Parameter2": (210, 570),  # רחוב+עיר
    "Parameter3": (100, 570),  # תז
    "Parameter4": (413, 553),  # מספר ריכוז
    "Parameter5": (269, 553),  # מיקום קלפי\ ריכוז
    "Parameter6": (148, 553),  # כתובת מיקום קלפי\ ריכוז
    "Parameter7": (0, 0),  # מספר קלפי
    "Parameter8": (330, 516),  # קלפיות החלפה

}

parameter_coordinates_sadran_hachvana = {
    "Parameter1": (357, 570),  # שם מלא
    # "Parameter2": (210, 570),  # רחוב+עיר
    # TODO chandeg 210=>205
    "Parameter2": (205, 570),  # רחוב+עיר
    "Parameter3": (100, 570),  # תז
    # "Parameter4": (406, 550),  # מספר ריכוז
    "Parameter4": (0, 0),  # מספר ריכוז
    "Parameter5": (300, 553),  # מיקום קלפי\ ריכוז
    "Parameter6": (165, 553),  # כתובת מיקום קלפי\ ריכוז
    "Parameter7": (0, 0),  # מספר קלפי
    "Parameter8": (0, 0),  # קלפיות החלפה

}


def generate_individual_pdf(row_data, output_dir):
    # pdf_filename_ini = f"{row[4] if row[4] is not None else ''}_{row[3] if row[3] is not None else ''}_{row[2] if row[2] is not None else ''}_{row[7] if row[7] is not None else ''}_{row[6] if row[6] is not None else ''}.pdf"
    pdf_filename_ini = f"{row[3] if row[3] is not None else ''}_{row[2] if row[2] is not None else ''}_{row[4] if row[4] is not None else ''}_{row[7] if row[7] is not None else ''}_{row[6] if row[6] is not None else ''}.pdf"
    pdf_filename_ini = re.sub('[<>:"\\|?*/]', "'", pdf_filename_ini)
    pdf_filename = output_dir + "/" + pdf_filename_ini

    template_canvas = canvas.Canvas(pdf_filename, pagesize=(2000, 2000))
    template_canvas.setFont('Hebrew_david', 11)
    font_size = 11
    parameter_index = 0  # Initialize parameter index

    if row_data[2] == "סדרן הכוונה" or row_data[2] == "סדרן מטה":
        for parameter, value in zip(row_data[3:], parameter_coordinates_sadran_hachvana.keys()):
            if parameter:
                print(f'parameter: {parameter}, value: {parameter_coordinates_sadran_hachvana[value]}')
                x_coord, y_coord = parameter_coordinates_sadran_hachvana[value]
                if x_coord == 0 and y_coord == 0:
                    continue  # Skip printing for (0, 0) coordinates

                if parameter_index == 1:
                    # TODO changed 24 => 26
                    if len(str(parameter)) > 26:
                        template_canvas.setFont('Hebrew_david', 9)
                        font_size = 9
                        x_coord = 200
                        if len(str(parameter)) < 30:
                            x_coord = calculate_num(x_coord, parameter, 8, 29)
                            # x_coord += 15
                    else:
                        template_canvas.setFont('Hebrew_david', 11)
                        x_coord, y_coord = parameter_coordinates_sadran_hachvana[value]
                        x_coord = calculate_num(x_coord, parameter, 4, 25)
                else:
                    template_canvas.setFont('Hebrew_david', 11)
                if value in ["Parameter1"]:
                    # x_coord = calculate_text_start(x_coord, 94, 11, parameter)
                    x_coord = calculate_num(x_coord, parameter, 5, 17)
                    # template_canvas.setFont('Hebrew_david', 16)
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    print("value1 - parameter: ", parameter)
                if value in ["Parameter2"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter3"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter4"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter5"]:
                    print("value1 - parameter: ", parameter)
                    x_coord = calculate_num(x_coord, parameter, 5, 21)
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter6"]:
                    print("value1 - parameter: ", parameter)
                    x_coord = calculate_num(x_coord, parameter, 5, 17)
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter7"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter8"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                parameter_index += 1

    else:
        if row_data[2] == 'מזכיר מחליף בריכוז' or row_data[2] == 'קולט' or row_data[2] == 'מזכיר שני':
            for parameter, value in zip(row_data[3:], parameter_coordinates_second_or_kolet.keys()):
                print(f'parameter: {parameter}, value: {parameter_coordinates_second_or_kolet[value]}')
                if parameter:
                    x_coord, y_coord = parameter_coordinates_second_or_kolet[value]
                    if x_coord == 0 and y_coord == 0:
                        continue  # Skip printing for (0, 0) coordinates
                    if parameter_index == 1:
                        # TODO changed 24 => 26
                        if len(str(parameter)) > 26:
                            template_canvas.setFont('Hebrew_david', 9)
                            font_size = 9
                            x_coord = 200
                            if len(str(parameter)) < 30:
                                x_coord = calculate_num(x_coord, parameter, 8, 29)
                        else:
                            template_canvas.setFont('Hebrew_david', 11)
                            x_coord, y_coord = parameter_coordinates_sadran_hachvana[value]
                            x_coord = calculate_num(x_coord, parameter, 4, 25)
                    else:
                        template_canvas.setFont('Hebrew_david', 11)
                    if value in ["Parameter1"]:
                        # template_canvas.setFont('Hebrew_david', 16)
                        x_coord = calculate_num(x_coord, parameter, 5, 17)
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter2"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter3"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter4"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter5"]:
                        x_coord = calculate_num(x_coord, parameter, 5, 22)
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter6"]:
                        x_coord = calculate_num(x_coord, parameter, 5, 18)
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter7"]:
                        template_canvas.drawString(x_coord + 5, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter8"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    parameter_index += 1

        else:
            # all the rest of the files
            print(f'{row_data[3]} {row_data[2]}')
            # Initialize an iterator for the parameters

            for parameter, value in zip(row_data[3:], parameter_coordinates_rest.keys()):
                print(f'parameter: {parameter}, value: {parameter_coordinates_rest[value]}')
                if parameter:
                    x_coord, y_coord = parameter_coordinates_rest[value]
                    if x_coord == 0 and y_coord == 0:
                        continue  # Skip printing for (0, 0) coordinates
                        # Check if we are processing the second parameter and adjust the font size

                    if parameter_index == 1:
                        # TODO changed 24 => 26
                        if len(str(parameter)) > 26:
                            template_canvas.setFont('Hebrew_david', 9)
                            x_coord = 200
                            if len(str(parameter)) < 30:
                                x_coord = calculate_num(x_coord, parameter, 8, 29)
                        else:
                            template_canvas.setFont('Hebrew_david', 11)
                            x_coord, y_coord = parameter_coordinates_sadran_hachvana[value]
                            x_coord = calculate_num(x_coord, parameter, 4, 25)
                    else:
                        template_canvas.setFont('Hebrew_david', 11)
                    if value in ["Parameter1"]:
                        # template_canvas.setFont('Hebrew_david', 16)
                        x_coord = calculate_num(x_coord, parameter, 5, 17)
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter2"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter3"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter4"]:
                        if not row_data[2] == 'מזכיר ראשון':
                            template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter5"]:
                        x_coord = calculate_num(x_coord, parameter, 5, 22)
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter6"]:
                        x_coord = calculate_num(x_coord, parameter, 5, 18)
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter7"]:
                        # kalpi num
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    if value in ["Parameter8"]:
                        template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                    # Increment the parameter index
                    parameter_index += 1

    template_canvas.save()
    return [row_data[1], row_data[2], pdf_filename]


# Load Excel data and generate individual PDFs
excel_file = 'input_data_new.xlsx'
workbook = load_workbook(excel_file)
sheet = workbook.active
# Initialize empty arrays for each category
ramat_hasharon = {
    "av_bait_vesadran": "רמש כתב מינוי לאב בית וסדרן.pdf",
    "sadran": "רמש כתב מינוי לסדרן.pdf",
    "sadran_hacvana": "רמש כתב מינוי לסדרן הכוונה.pdf",
    "mazkir": "רמש כתב מינוי מזכיר.pdf",
    "mazkir_mahlif": "רמש כתב מינוי מזכיר מחליף.pdf",
    "kolet": "רמש כתב מינוי קולט.pdf",
    "sadran_mate": "רמש כתב מינוי לסדרן מטה.pdf",
    "mazkir_movil": "רמש כתב מינוי מזכיר מוביל.pdf",
}

bat_yam = {
    "av_bait_vesadran": "בת ים כתב מינוי לאב בית וסדרן.pdf",
    "sadran": "בת ים כתב מינוי לסדרן.pdf",
    "sadran_hacvana": "בת ים כתב מינוי לסדרן הכוונה.pdf",
    "mazkir": "בת ים כתב מינוי מזכיר.pdf",
    "mazkir_mahlif": "בת ים כתב מינוי מזכיר שני.pdf",
    "kolet": "בת ים כתב מינוי קולט.pdf",
    "sadran_mate": "בת ים כתב מינוי לסדרן מטה.pdf",

}
rishon = {
    "av_bait_vesadran": "ראשון כתב מינוי לאב בית וסדרן.pdf",
    "sadran": "ראשון כתב מינוי לסדרן.pdf",
    "sadran_hacvana": "ראשון כתב מינוי לסדרן הכוונה.pdf",
    "mazkir": "ראשון כתב מינוי מזכיר.pdf",
    "mazkir_mahlif": "ראשון כתב מינוי מזכיר שני.pdf",
    "kolet": "ראשון כתב מינוי קולט.pdf",
    "sadran_mate": "ראשון כתב מינוי לסדרן מטה.pdf",
}
kfar_saba = {
    "av_bait_vesadran": "כפס כתב מינוי לאב בית וסדרן.pdf",
    "sadran": "כפס כתב מינוי לסדרן.pdf",
    "sadran_hacvana": "כפס כתב מינוי לסדרן הכוונה.pdf",
    "mazkir": "כפס כתב מינוי מזכיר.pdf",
    "mazkir_mahlif": "כפס כתב מינוי מזכיר שני.pdf",
    "kolet": "כפס כתב מינוי קולט.pdf",
    "sadran_mate": "כפס כתב מינוי לסדרן מטה.pdf",
}
output_directory = 'miniu'

if not os.path.exists(output_directory):
    os.makedirs(output_directory)
# Define the initial template file
initial_template_file_path = ''
for row in sheet.iter_rows(min_row=2, max_col=11, values_only=True):
    [city, job_description, pdf_filename] = generate_individual_pdf(row, output_directory)
    try:
        # Define the template file based on the city and job description
        if city == 'רמת השרון':
            if job_description == 'מזכיר ראשון' or job_description == 'מזכיר שני':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir"]}'
            elif job_description == 'מזכיר מחליף בריכוז':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran_hacvana"]}'
            elif job_description == 'סדרן מטה':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran_mate"]}'
            elif job_description == 'אב בית סדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{ramat_hasharon["kolet"]}'
            elif job_description == 'מזכיר מוביל':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir_movil"]}'
            else:
                continue

        if city == 'בת ים':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{bat_yam["mazkir"]}'
            elif job_description == 'מזכיר שני':
                template_file_path = f'pdf_files/{bat_yam["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{bat_yam["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{bat_yam["sadran_hacvana"]}'
            elif job_description == 'סדרן מטה':
                template_file_path = f'pdf_files/{bat_yam["sadran_mate"]}'
            elif job_description == 'אב בית סדרן':
                template_file_path = f'pdf_files/{bat_yam["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{bat_yam["kolet"]}'
            else:
                continue
        if city == 'כפר סבא':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{kfar_saba["mazkir"]}'
            elif job_description == 'מזכיר שני':
                template_file_path = f'pdf_files/{kfar_saba["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{kfar_saba["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{kfar_saba["sadran_hacvana"]}'
            elif job_description == 'סדרן מטה':
                template_file_path = f'pdf_files/{kfar_saba["sadran_mate"]}'
            elif job_description == 'אב בית סדרן':
                template_file_path = f'pdf_files/{kfar_saba["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{kfar_saba["kolet"]}'
            else:
                continue

        if city == 'ראשון לציון':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{rishon["mazkir"]}'
            elif job_description == 'מזכיר שני':
                template_file_path = f'pdf_files/{rishon["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{rishon["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{rishon["sadran_hacvana"]}'
            elif job_description == 'סדרן מטה':
                template_file_path = f'pdf_files/{rishon["sadran_mate"]}'
            elif job_description == 'אב בית סדרן':
                template_file_path = f'pdf_files/{rishon["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{rishon["kolet"]}'
            else:
                continue
        # Open the PDF files in binary mode
        template_file = open(template_file_path, 'rb')
        data_file = open(pdf_filename, 'rb')

        # Create PDF reader objects for the files
        template_reader = PyPDF2.PdfReader(template_file)
        data_reader = PyPDF2.PdfReader(data_file)

        # Create a PDF writer object to write the merged output
        merged_writer = PyPDF2.PdfWriter()

        # Get the first page of the template PDF
        template_page = template_reader.pages[0]

        # Get the first page of the data PDF
        data_page = data_reader.pages[0]

        # Merge the data page onto the template page
        template_page.merge_page(data_page)

        # Add the merged page to the output PDF
        merged_writer.add_page(template_page)

        # Save the output PDF file
        # final_output_filename_ini = f"{row[4]}_{row[3]}_{row[2]}_{row[7]}_{row[6]}.pdf"
        # print(f' XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX - {row[4]}')
        # final_output_filename_ini = f"{row[4] if row[4] is not None else ''}_{row[3] if row[3] is not None else ''}_{row[2] if row[2] is not None else ''}_{row[7] if row[7] is not None else ''}_{row[6] if row[6] is not None else ''}.pdf"
        final_output_filename_ini = f"{row[3] if row[3] is not None else ''}_{row[2] if row[2] is not None else ''}_{row[4] if row[4] is not None else ''}_{row[7] if row[7] is not None else ''}_{row[6] if row[6] is not None else ''}.pdf"

        final_output_filename = re.sub('[<>:"\\|?*/]', "'", final_output_filename_ini)
        final_output_filename = f"{output_directory}/{final_output_filename}"

        with open(final_output_filename, 'wb') as output_file:
            merged_writer.write(output_file)

    except Exception as e:
        print(f"Error processing row: {str(e)}")

    finally:
        # Close the input and output files if they were successfully opened
        if 'template_file' in locals() and template_file is not None:
            template_file.close()
        if 'data_file' in locals() and data_file is not None:
            data_file.close()
        print(f'{pdf_filename} was created')
