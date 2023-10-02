import PyPDF2
import os

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook

# pdfmetrics.registerFont(TTFont('Hebrew', 'Assistant-VariableFont_wght.ttf'))
pdfmetrics.registerFont(TTFont('Hebrew', 'arial-hebrew-bold.ttf'))
pdfmetrics.registerFont(TTFont('Hebrew_david', 'DavidLibre-Bold.ttf'))


# def reverse_slicing(s):
#     if isinstance(s, str):
#         return s[::-1]
#     return str(s)  # Convert non-string values to strings
#
def reverse_slicing(s):
    if isinstance(s, str):
        result = ""
        non_digit_chars = ""
        digit_chars = ""

        for char in s:
            if char.isdigit():
                if non_digit_chars:
                    result = non_digit_chars[::-1] + result
                    non_digit_chars = ""
                digit_chars += char
            else:
                if digit_chars:
                    result = digit_chars + result
                    digit_chars = ""
                non_digit_chars += char

        result = digit_chars + result
        result = non_digit_chars[::-1] + result

        return result
    return str(s)


# def reverse_non_digits_with_digits(s):
# def reverse_slicing(s):
#     if isinstance(s, str):
#         non_digit_chars = [char for char in s if not char.isdigit()]
#         reversed_non_digit = ''.join(non_digit_chars[::-1])
#
#         result = ""
#         digit_chars = ""
#
#         for char in s:
#             if char.isdigit():
#                 digit_chars += char
#             else:
#                 if len(reversed_non_digit) > 0:
#                     result += reversed_non_digit[0]
#                     reversed_non_digit = reversed_non_digit[1:]
#
#         result = digit_chars + result
#
#         return result
#     return str(s)  # Convert non-string values to strings


# Define a mapping of parameter names/indices to x, y coordinates
parameter_coordinates = {
    "Parameter1": (165, 712),  # עיר כותבת
    "Parameter2": (230, 620),  # כתב מינוי
    "Parameter3": (300, 528),  # עיר
    "Parameter4": (0, 510),  # שם מלא
    "Parameter5": (-100, -100),  # רחוב+עיר
    "Parameter6": (-100, -100),  # תז
    "Parameter7": (0, 488),  # לקלפי\מקום ריכוז
    "Parameter8": (-100, -100),  # מספר
    "Parameter9": (-100, -100),  # כתובת מלאה + עיר
    "Parameter10": (400, 450),  # עיר.
    "Parameter11": (0, 466),  # קלפיות החלפה
    "Parameter12": (243, 293),  # מנהל בחירות
    "Parameter13": (265, 310),  # עיר ושנה
}
PAGE_WIDTH, PAGE_HEIGHT = letter
CENTER_X = PAGE_WIDTH / 2


def generate_individual_pdf(row_data, output_dir):
    pdf_filename = f"{output_dir}/כתב מינוי_{row_data[4]}.pdf"
    template_canvas = canvas.Canvas(pdf_filename, pagesize=(2000, 2000))
    template_canvas.setFont('Hebrew', 10)
    first_row_flag = 0
    second_row_flag = 0
    third_row_flag = 0

    for parameter, value in zip(row_data[1:], parameter_coordinates.keys()):
        if parameter:
            x_coord, y_coord = parameter_coordinates[value]
            if value in ["Parameter1"]:
                template_canvas.setFont('Hebrew_david', 16)
                template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
            else:
                if value in ["Parameter2"]:
                    template_canvas.setFont('Hebrew', 20)
                    text_width = template_canvas.stringWidth(parameter, 'Hebrew', 20)
                    new_x_coord = CENTER_X - (text_width / 2)
                    template_canvas.drawString(new_x_coord, y_coord, reverse_slicing(parameter))
                else:
                    if value in ["Parameter4", "Parameter5", "Parameter6"] and not first_row_flag:
                        first_row_flag = 1
                        concat_line = "אני ממנה את " + str(row_data[row_data.index(parameter)]) + " מרחוב " + str(
                            row_data[row_data.index(parameter) + 1]) + "," + " מס' ת.ז " + str(
                            row_data[row_data.index(parameter) + 2])
                        text_width = template_canvas.stringWidth(parameter, 'Hebrew', 12)
                        template_canvas.setFont('Hebrew', 12)
                        new_x_coord = CENTER_X - (text_width / 2) - 150
                        template_canvas.drawString(new_x_coord, y_coord, reverse_slicing(concat_line))
                        print(concat_line)
                        print(x_coord, y_coord)
                    else:  # Skip drawing the individual parameters
                        if value in ["Parameter7", "Parameter8", "Parameter9"] and not second_row_flag:
                            second_row_flag = 1
                            # concat_line = "לקלפי מס' " + str(
                            concat_line = str(
                                row_data[row_data.index(parameter)]) + " " + str(
                                row_data[row_data.index(parameter) + 1]) + " ב" + str(
                                row_data[row_data.index(parameter) + 2])
                            text_width = template_canvas.stringWidth(parameter, 'Hebrew', 12)
                            template_canvas.setFont('Hebrew', 12)
                            new_x_coord = CENTER_X - (text_width / 2) - 150
                            template_canvas.drawString(new_x_coord, y_coord, reverse_slicing(concat_line))
                            print(concat_line)
                            print(x_coord, y_coord)
                        else:
                            if value in ["Parameter11"] and not third_row_flag:
                                third_row_flag = 1
                                concat_line = str(row_data[row_data.index(parameter)]) + reverse_slicing(
                                    "לקלפיות החלפה מספר: ")

                                template_canvas.setFont('Hebrew', 12)
                                text_width = template_canvas.stringWidth(parameter, 'Hebrew', 12)
                                new_x_coord = CENTER_X - (text_width / 2) - 50
                                # template_canvas.drawString(new_x_coord, y_coord, reverse_slicing(concat_line))
                                template_canvas.drawString(new_x_coord, y_coord, concat_line)
                                print(concat_line)
                                print(x_coord, y_coord)
                            else:
                                template_canvas.setFont('Hebrew', 12)
                                template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                            # print(reverse_slicing(parameter))

    template_canvas.save()
    return pdf_filename


# Load Excel data and generate individual PDFs
excel_file = 'input_data.xlsx'
workbook = load_workbook(excel_file)
sheet = workbook.active

output_directory = 'miniu'
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

for row in sheet.iter_rows(min_row=2, max_col=17, values_only=True):
    pdf_filename = generate_individual_pdf(row, output_directory)

    # Open the PDF files in binary mode
    template_file = open('pdf_files/minui_template.pdf', 'rb')
    data_file = open(pdf_filename, 'rb')

    # Create PDF reader objects for the files
    template_reader = PyPDF2.PdfReader(template_file)
    data_reader = PyPDF2.PdfReader(data_file)

    # Create a PDF writer object to write the merged output
    merged_writer = PyPDF2.PdfWriter()

    # Get the first page of the template PDF
    template_page = template_reader.pages[0]

    # Get the first page of the data PDF
    data_page = data_reader.pages[0]

    # Merge the data page onto the template page
    template_page.merge_page(data_page)

    # Add the merged page to the output PDF
    merged_writer.add_page(template_page)

    # Save the output PDF file
    final_output_filename = f"{output_directory}/כתב מינוי_{row[4]}.pdf"
    with open(final_output_filename, 'wb') as output_file:
        merged_writer.write(output_file)

    # Close the input and output files
    template_file.close()
    data_file.close()
