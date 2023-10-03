import re

import PyPDF2
import os

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook

# pdfmetrics.registerFont(TTFont('Hebrew', 'Assistant-VariableFont_wght.ttf'))
pdfmetrics.registerFont(TTFont('Hebrew', 'arial-hebrew-bold.ttf'))
pdfmetrics.registerFont(TTFont('Hebrew_david', 'DavidLibre-Bold.ttf'))


# def reverse_slicing(s):
#     if isinstance(s, str):
#         return s[::-1]
#     return str(s)  # Convert non-string values to strings
# def reverse_slicing(s):
#     if isinstance(s, str):
#         result = ""
#         non_digit_chars = ""
#         digit_chars = ""
#
#         for char in s:
#             if char.isdigit():
#                 if non_digit_chars:
#                     result = non_digit_chars[::-1] + result
#                     non_digit_chars = ""
#                 digit_chars += char
#             else:
#                 if digit_chars:
#                     result = digit_chars + result
#                     digit_chars = ""
#                 non_digit_chars += char
#
#         result = digit_chars + result
#         result = non_digit_chars[::-1] + result
#
#         return result
#     return str(s)
def reverse_slicing(s):
    s = str(s)
    s = list(s)
    i, j = 0, len(s) - 1
    while i < j:
        if (s[i].isdigit() and s[j].isdigit()) or (s[i] == '(' and s[j] == ')'):
            s[i], s[j] = s[j], s[i]
            i += 1
            j -= 1
        elif s[i].isdigit() or s[i] == '(':
            j -= 1
        elif s[j].isdigit() or s[j] == ')':
            i += 1
        else:
            s[i], s[j] = s[j], s[i]
            i += 1
            j -= 1
    return ''.join(s)


# Define a mapping of parameter names/indices to x, y coordinates
parameter_coordinates_rest = {

    "Parameter1": (357, 567),  # שם מלא
    "Parameter2": (213, 567),  # רחוב+עיר
    "Parameter3": (100, 567),  # תז
    "Parameter4": (413, 550),  # מספר קלפי\ריכוז
    "Parameter5": (269, 550),  # מיקום קלפי\ ריכוז
    "Parameter6": (148, 550),  # כתובת מיקום קלפי\ ריכוז
    "Parameter7": (330, 514),  # קלפיות החלפה

}

parameter_coordinates_sadran_hachvana = {
    "Parameter1": (357, 567),  # שם מלא
    "Parameter2": (210, 567),  # רחוב+עיר
    "Parameter3": (100, 567),  # תז
    "Parameter4": (406, 550),  # מספר קלפי\ריכוז
    "Parameter5": (285, 550),  # מיקום קלפי\ ריכוז
    "Parameter6": (165, 550),  # כתובת מיקום קלפי\ ריכוז
    "Parameter7": (330, 514),  # קלפיות החלפה

}


def generate_individual_pdf(row_data, output_dir):
    # pdf_filename = f"{output_dir}/כתב מינוי_{row[3]}_{row[2]}.pdf"
    pdf_filename_ini = f"{output_dir}/{row[3]}_{row[2]}_{row[7]}_{row[6]}.pdf"
    # pdf_filename = final_output_filename.replace('"', ''')
    pdf_filename = re.sub('[<>:"\\|?*]', "'", pdf_filename_ini)

    # pdf_filename = f"{output_dir}/{row[3]}_{row[2]}_{row[7]}_{row[6]}.pdf"
    template_canvas = canvas.Canvas(pdf_filename, pagesize=(2000, 2000))
    # template_canvas.setFont('Hebrew', 10)

    # print("------", row_data[2])
    if row_data[2] == "סדרן הכוונה":
        for parameter, value in zip(row_data[3:], parameter_coordinates_sadran_hachvana.keys()):
            if parameter:
                x_coord, y_coord = parameter_coordinates_sadran_hachvana[value]
                template_canvas.setFont('Hebrew_david', 11)
                if value in ["Parameter1"]:
                    # template_canvas.setFont('Hebrew_david', 16)
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter2"]:
                    # template_canvas.setFont('Hebrew', 20)
                    # text_width = template_canvas.stringWidth(parameter, 'Hebrew', 20)
                    # new_x_coord = CENTER_X - (text_width / 2)
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter3"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                # if value in ["Parameter4"]:
                #     template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter5"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter6"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter7"]:
                    template_canvas.drawString(x_coord, y_coord, parameter)
    else:
        # if row_data[1] == 'מזכיר קלפי חלופי':
        for parameter, value in zip(row_data[3:], parameter_coordinates_rest.keys()):
            if parameter:
                x_coord, y_coord = parameter_coordinates_rest[value]
                template_canvas.setFont('Hebrew_david', 11)
                if value in ["Parameter1"]:
                    # template_canvas.setFont('Hebrew_david', 16)
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter2"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter3"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter4"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter5"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter6"]:
                    template_canvas.drawString(x_coord, y_coord, reverse_slicing(parameter))
                if value in ["Parameter7"]:
                    template_canvas.drawString(x_coord, y_coord, parameter)

    template_canvas.save()
    return [row_data[1], row_data[2], pdf_filename]
    # return pdf_filename


# Load Excel data and generate individual PDFs
excel_file = 'input_data_new.xlsx'
workbook = load_workbook(excel_file)
sheet = workbook.active
# Initialize empty arrays for each category
ramat_hasharon = {
    "av_bait_vesadran": "רמש כתב מינוי לאב בית וסדרן.pdf",
    "sadran": "רמש כתב מינוי לסדרן.pdf",
    "sadran_hacvana": "רמש כתב מינוי לסדרן הכוונה.pdf",
    "mazkir": "רמש כתב מינוי מזכיר.pdf",
    "mazkir_mahlif": "רמש כתב מינוי מזכיר מחליף.pdf",
    "kolet": "רמש כתב מינוי קולט.pdf",
}

bat_yam = {
    "av_bait_vesadran": "בת ים כתב מינוי לאב בית וסדרן.pdf",
    "sadran": "בת ים כתב מינוי לסדרן.pdf",
    "sadran_hacvana": "בת ים כתב מינוי לסדרן הכוונה.pdf",
    "mazkir": "רמש כתב מינוי מזכיר.pdf",
    "mazkir_mahlif": "בת ים כתב מינוי מזכיר מחליף.pdf",
    "kolet": "בת ים כתב מינוי קולט.pdf",
}
rishon = []
kfar_saba = []
output_directory = 'miniu'

if not os.path.exists(output_directory):
    os.makedirs(output_directory)
# Define the initial template file
initial_template_file_path = ''
for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
    [city, job_description, pdf_filename] = generate_individual_pdf(row, output_directory)

    try:
        # Define the template file based on the city and job description
        if city == 'רמת השרון':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir"]}'
            elif job_description == 'מזכיר מחליף':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran_hacvana"]}'
            elif job_description == 'אב בית וסדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{ramat_hasharon["kolet"]}'
            else:
                continue

        if city == 'בת ים':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir"]}'
            elif job_description == 'מזכיר מחליף':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran_hacvana"]}'
            elif job_description == 'אב בית וסדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{ramat_hasharon["kolet"]}'
            else:
                continue
        # todo לשנטת שם סביבה
        if city == 'כפר סבא':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir"]}'
            elif job_description == 'מזכיר מחליף':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran_hacvana"]}'
            elif job_description == 'אב בית וסדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{ramat_hasharon["kolet"]}'
            else:
                continue

        if city == 'ראשון לציון':
            if job_description == 'מזכיר':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir"]}'
            elif job_description == 'מזכיר מחליף':
                template_file_path = f'pdf_files/{ramat_hasharon["mazkir_mahlif"]}'
            elif job_description == 'סדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran"]}'
            elif job_description == 'סדרן הכוונה':
                template_file_path = f'pdf_files/{ramat_hasharon["sadran_hacvana"]}'
            elif job_description == 'אב בית וסדרן':
                template_file_path = f'pdf_files/{ramat_hasharon["av_bait_vesadran"]}'
            elif job_description == 'קולט':
                template_file_path = f'pdf_files/{ramat_hasharon["kolet"]}'
            else:
                continue
        #         template_file_path = initial_template_file_path  # Use the initial template file as a default
        # else:
        #     template_file_path = initial_template_file_path  # Use the initial template file as a default

        # Open the PDF files in binary mode
        template_file = open(template_file_path, 'rb')
        data_file = open(pdf_filename, 'rb')

        # Create PDF reader objects for the files
        template_reader = PyPDF2.PdfReader(template_file)
        data_reader = PyPDF2.PdfReader(data_file)

        # Create a PDF writer object to write the merged output
        merged_writer = PyPDF2.PdfWriter()

        # Get the first page of the template PDF
        template_page = template_reader.pages[0]

        # Get the first page of the data PDF
        data_page = data_reader.pages[0]

        # Merge the data page onto the template page
        template_page.merge_page(data_page)

        # Add the merged page to the output PDF
        merged_writer.add_page(template_page)

        # Save the output PDF file
        # final_output_filename = f"{output_directory}/כתב מינוי_{row[3]}_{row[2]}.pdf"
        final_output_filename_ini = f"{output_directory}/{row[3]}_{row[2]}_{row[7]}_{row[6]}.pdf"
        final_output_filename = re.sub('[<>:"\\|?*]', "'", final_output_filename_ini)

        with open(final_output_filename, 'wb') as output_file:
            merged_writer.write(output_file)

    except Exception as e:
        print(f"Error processing row: {str(e)}")

    finally:
        # Close the input and output files if they were successfully opened
        if 'template_file' in locals() and template_file is not None:
            template_file.close()
        if 'data_file' in locals() and data_file is not None:
            data_file.close()
        print(f'{pdf_filename} was created')
